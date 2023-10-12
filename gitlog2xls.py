# For logging during interpreting
import logging
# For convert dateformat from parsed string
from datetime import datetime 
# For unsing Excel workbook
import openpyxl as op
# For using Excel workbook with styles
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


#logging.basicConfig(level=logging.DEBUG,filename="gitlogparser.log",datefmt = '%Y-%m-%d %H:%M:%S',format = '%(asctime)s | %(levelname)s | %(message)s')
logging.basicConfig(level=logging.DEBUG,datefmt = '%Y-%m-%d %H:%M:%S', format = '%(asctime)s | %(levelname)s | %(message)s')
fp = open('git_commit.log', 'rt', encoding='utf-8')
lines = fp.readlines()
lineLen = len(lines)
commitCnt = 1
lineNo = 0
writeToXls = False
cntNewLine = 0
noMsgLine = 0
commitMsg = []
commitAuthor = []
commitAuthorEmail = []
commitID = []
commitDate =  []

wb = op.Workbook()
ws = wb.create_sheet('GIT LOG',0)
#wsTest = wb['Sheet']
#wsTest.title='Test'
ws["A1"].value = "No"
ws["B1"].value = "Date"
ws["C1"].value = "Commit"
ws["D1"].value = "Author"
ws["E1"].value = "E-mail"
ws["F1"].value = "Messages"
titleFont = Font(name='Arial', bold=True, size=12, italic=True, color="00FF0000")
titleFill = PatternFill(fill_type='solid', fgColor="70C0C0")
msgCellAlign= Alignment(horizontal='left',vertical='center')
msgCellFill = PatternFill(fill_type='solid', fgColor="A0A0A0")
defaultCellAlign = Alignment(horizontal='center', vertical='center')
defaultBorder =  Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'), 
                        top=Side(border_style='thin', color='000000'), 
                        bottom=Side(border_style='thin', color='000000'))

for i in range(1, 7):
    ws.cell(row=1, column=i).font = titleFont
    ws.cell(row=1, column=i).fill = titleFill 
    ws.cell(row=1, column=i).alignment = defaultCellAlign
    ws.cell(row=1, column=i).border = defaultBorder

ws.freeze_panes = "A2"
ws.column_dimensions['A'].width=7
ws.column_dimensions['B'].width=25
ws.column_dimensions['C'].width=45
ws.column_dimensions['D'].width=25
ws.column_dimensions['E'].width=25
ws.column_dimensions['F'].width=150
ws.row_dimensions[1].height = 30

while lineNo < lineLen:
    itr = lines[lineNo]
    if itr == "\n":
        cntNewLine += 1
        if cntNewLine == 2:
            writeToXls = True
    else:
        if(itr.startswith('commit ')):
            commitID = itr.split("commit ",1)[1].split("\n",1)[0]
        elif(itr.startswith("Author:")):
            splitData = itr.split("Author:",1)[1].split(" ",1)[1]
            commitAuthor = splitData.split("<",1)[0]
            commitAuthorEmail = splitData.split("<",1)[1].split(">",1)[0]
        elif(itr.startswith("Date:")):
            commitDate = datetime.strptime(itr.split("Date:",1)[1].split("\n",1)[0], '   %a %b %d %H:%M:%S %Y %z').strftime('%Y-%m-%d %H:%M:%S')
        else:
            if(cntNewLine == 1):
                noMsgLine += 1
                commitMsg.append(itr.split("    ",1)[1])

    if writeToXls == True:
        commitMsg = ''.join(commitMsg)
        commitMsg = commitMsg.rstrip('\r\n')
        logging.debug('===== Commit No [%d]=====================', commitCnt)
        logging.debug('Commit ID : %s', commitID)
        logging.debug('Author    : %s', commitAuthor)
        logging.debug('E-mail    : %s', commitAuthorEmail)
        logging.debug('Date      : %s', commitDate)
        logging.debug('Commit Msg: %s', commitMsg)

        # Fill parsed string into cell
        ws.cell(row=commitCnt+1, column=1).value = commitCnt
        ws.cell(row=commitCnt+1, column=2).value = commitDate
        ws.cell(row=commitCnt+1, column=3).value = commitID
        ws.cell(row=commitCnt+1, column=4).value = commitAuthor
        ws.cell(row=commitCnt+1, column=5).value = commitAuthorEmail
        ws.cell(row=commitCnt+1, column=6).value = str(commitMsg)

        # Adopt style as condition
        for i in range(1, 7):
            if "version history" in commitMsg:
                ws.cell(row=commitCnt+1, column=i).fill = msgCellFill
            ws.cell(row=commitCnt+1, column=i).border = defaultBorder
            ws.cell(row=commitCnt+1, column=i).alignment = defaultCellAlign
            if(i==6):
                if(noMsgLine > 2):
                    ws.cell(row=commitCnt+1, column=6).alignment = Alignment(wrap_text=True, vertical='center')
                else:
                    ws.cell(row=commitCnt+1, column=6).alignment = msgCellAlign

        commitCnt += 1
        commitMsg = []
        writeToXls = False
        cntNewLine = 0
        noMsgLine = 0

    lineNo = lineNo + 1

fp.close()
wb.save(r"D:\PyWorks\GitLogParser\gitlogparser.xlsx")
